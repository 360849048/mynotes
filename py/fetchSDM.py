from urllib import request, parse
from http import cookiejar
from bs4 import BeautifulSoup
import openpyxl
from queue import Queue
import re
import threading


cookie = cookiejar.CookieJar()
# 获取__VIEWSTATE和__EVENTVALIDATION数据
def _getLoginInfo():
    url = 'http://192.168.100.179/HTMSOFT/login.aspx'
    header = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        # 'Content-Length': 294,
        'Content-Type': 'application/x-www-form-urlencoded',
        'Host': '192.168.100.179',
        'Origin': 'http://192.168.100.179',
        'Refer': 'http://192.168.100.179/HTMSOFT/login.aspx',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36'
    }
    __VIEWSTATE = ''
    __EVENTVALIDATION = ''
    try:
        req = request.Request(url=url, headers=header)
        page = request.urlopen(req).read()
        soup = BeautifulSoup(page, 'html.parser')
        __VIEWSTATE = soup.find("input", {"id": "__VIEWSTATE"}).attrs["value"]
        __EVENTVALIDATION = soup.find("input", {"id": "__EVENTVALIDATION"}).attrs["value"]
    except Exception as e:
        print(e)
    return(__VIEWSTATE, __EVENTVALIDATION)

# 下面代码用来获取cookie
def login(username, password):
    __VIEWSTATE, __EVENTVALIDATION = _getLoginInfo()
    url = 'http://192.168.100.179/HTMSOFT/login.aspx'
    data = {'tbUserName': username,
            'tbPassword': password,
            'btnLogin.x': '0',
            'btnLogin.y': '0',
            '__VIEWSTATE': __VIEWSTATE,
            '__EVENTVALIDATION': __EVENTVALIDATION
    }
    header = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        # 'Content-Length': 294,
        'Content-Type': 'application/x-www-form-urlencoded',
        'Host': '192.168.100.179',
        'Origin': 'http://192.168.100.179',
        'Refer': 'http://192.168.100.179/HTMSOFT/login.aspx',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36'
    }

    data = parse.urlencode(data).encode()
    req = request.Request(url=url, data=data, headers=header)
    handler = request.HTTPCookieProcessor(cookie)
    opener = request.build_opener(handler)
    page = opener.open(req).read()

    # for item in cookie:
        # print(item.name, '\t', item.value)
    # print(page.decode())

def fetchInfo(order):
    url = 'http://192.168.100.179/HTMSOFT/TasksQuery.aspx'

    data = {'Order_no': str(order),
            'Status': '',
            'BtnSave': '查询',
            'order': 'plan_finish_date',
            '__VIEWSTATE': '/wEPDwUJMzQ2MDAxMTUwD2QWAgIDD2QWJgIBDxAPFgIeC18hRGF0YUJvdW5kZ2QQFQkG6K6i5Y2VBuWFtuS7lgbllK7lkI4M546w5Zy65pSv5oyBDOaWsOWTgeW8gOWPkQzml6XluLjnlJ/kuqcP55Sf5Lqn5pu05pS55Y2VEuWKn+iDveaUuei/m+aPkOWNhwAVCQborqLljZUG5YW25LuWBuWUruWQjgznjrDlnLrmlK/mjIEM5paw5ZOB5byA5Y+RDOaXpeW4uOeUn+S6pw/nlJ/kuqfmm7TmlLnljZUS5Yqf6IO95pS56L+b5o+Q5Y2HABQrAwlnZ2dnZ2dnZ2dkZAIDDxAPFgIfAGdkEBUCCFNJR01BVEVLABUCCFNJR01BVEVLABQrAwJnZ2RkAhsPEA8WAh8AZ2QQFQkD5pawDOato+WcqOWkhOeQhg/nvJbnoIHlt7LkuIrkvKAOSU/ooajlt7LkuIrkvKAV56iL5bqP5a6M5oiQ5b6F5LiK5LygCeW3suS4iuS8oAblrozmiJAJ5pyq5a6M5oiQABUJA+aWsAzmraPlnKjlpITnkIYP57yW56CB5bey5LiK5LygDklP6KGo5bey5LiK5LygFeeoi+W6j+WujOaIkOW+heS4iuS8oAnlt7LkuIrkvKAG5a6M5oiQCeacquWujOaIkAAUKwMJZ2dnZ2dnZ2dnZGQCMw8QZGQWAWZkAjUPDxYCHgRUZXh0BQExZGQCNw8PFgIfAQUBMWRkAjkPDxYCHwEFATFkZAI7Dw8WAh4HRW5hYmxlZGhkZAI9Dw8WAh8CaGRkAj8PDxYCHwJoZGQCQQ8PFgIfAmhkZAJHDzwrAAsBAA8WCh4QQ3VycmVudFBhZ2VJbmRleGYeCERhdGFLZXlzFgAeC18hSXRlbUNvdW50AgEeCVBhZ2VDb3VudAIBHhVfIURhdGFTb3VyY2VJdGVtQ291bnQCAWQWAmYPZBYCAgIPZBYiZg8PFgIfAQUFODE5NDFkZAIBDw8WAh8BBQborqLljZVkZAICDw8WAh8BBQo2MDcwMDAyNzU2ZGQCAw8PFgIfAQUKMjAyMC0wMDkwOGRkAgQPDxYCHwEFEuiLseWbve+8jOayiOa1t+aYjmRkAgUPDxYCHwEFClpFMTIwMOKFoVNkZAIGDw8WAh8BBZ0C6Z2e5qCHCumpseWKqOWZqO+8mkY2CuWuieWFqOagh+WHhu+8mkNFCuWuieWFqOe7p+eUteWZqOeoi+W6j++8mklJc19FdSZVTF8xLjAuMCDlkowgRXVyb21hcDczLjIuMi4yCuadoeasvu+8mjEuMue7hOS4reWtkCAgIAoyLuasp+inhDEy5py65qKw5omL5o6l5Y+jICAgICAgICAgICAgICAgICAgCjMu6Imy5q+N5py65L+h5Y+35o6l5Y+jCjQuRVU3M+WkluWbtOWuieWFqOiuvuWkh+eUteawlOaOpeWPoyAgCjUu5qih5YW36aG25p2/6YCA5Zue5a6J5YWo56Gu6K6k5byA5YWz5o6l5Y+jICAgCjYuWmVzZGQCBw8PFgIfAQUIU0lHTUFURUtkZAIIDw8WAh8BBQnlupTlv5fls7BkZAIJDw8WAh8BBRXlj7bmsZ/luIXvvJvpg5HlrofosapkZAIKDw8WAh8BBQoyMDIwLTAzLTIwZGQCCw8PFgIfAQUD5pawZGQCDA8PFgIfAQUKMjAyMC0wMS0xOWRkAg0PDxYCHwEFAiAgZGQCDg8PFgIfAQUKMjAyMC0wMS0xOWRkAg8PDxYCHwEFDzIwMjAxNTAxMjAwMjI0N2RkAhAPZBYCZg8VAh5qYXZhc2NyaXB0OlNob3dTaW5nbGUoJzgxOTQxJykG5p+l55yLZAJJDw8WAh8BBQExZGQCSw8PFgIfAQUBMWRkAk0PDxYCHwEFATFkZAJPDw8WAh8CaGRkAlEPDxYCHwJoZGQCUw8PFgIfAmhkZAJVDw8WAh8CaGRkZHKfkKvaW6Ra85rsUHCl427DrAF3uV+pJh9xawwRIBon',
            '__EVENTVALIDATION': '/wEdAC431nYSIRbEtn5cSYYnU92hCfNaP5jZKyEnVrZbcu2+1Ub0VyDGgL5EPJkSP4SYi0/BKTvVg7Q1eAgRue8xqVfrujCm9FpYuod1rGg8nbq2Ys5zxM20UyWvnfavnJ+W7v6UaPQJD4t/P+rzS4p6fb/EQOJrK3lqu3QvkM6rU7zYOLaBSf5r+BZFG4w2vMyswwvYm4vsIqZD5z3NTwKvJfA8AL6H5j9QO4IeE+f3Zn3NaLXOLTbI5tWsrABBRXU2P9HovHoaMuQ7H4gMGdXUC0Cffcg65/VoH1HIvnctTUon5sZvAYp6n6LMM8LjVj1mxdFNhOPigeiu5DtF73bcLfPOoYh1fme5vmXh6pfAx4mVx8mtTaWgLWLoc5T7w4F5a6DQz6IKy5lyKw6CI5kTmnDn5gssTb5LTwIdyomz0bzj6IBBxVZXvdpcvUM3IET4aU3mcYGUd8MaHbmSPFNsLnDpRLcTUxcQggqTi7JfpPqzbAdpONkuFMkNGHf6fkLaRYcGq5+M7jl8ZFiYcXpgHCGTZexkljPdhShivMEH5eX2cjXbmSCuTYdtIHVTf4pwpsMSBL+LYrRnqEWsQOVh8btgF5w17e8IL156Gti+zNyEbbGZZpFVGkjHY8wb50Iw43dRJ3j5y2c9YXEpONWJQ6IARyrzmTjBuz0uRwuaCMs3qjda+ygl9rVl0o+bbUVXdhmA8Qhg4vM322xO3fysxs68T1tIlPWz3K7nlepLZtWhSYrNQZFkwR7WkpE9C23JRjyZb/KcGSSQDbEj3mKTxWzgESCFkFW/RuhzY1oLb/NUVEYWjLov5K6FKlBe92FmdG4o3AsueJVRtDGYyKfrpI37lVtu17WB70CjHkb72akyaTtmUrwMm8ZW4EkkSG3gNqXH9pRb4Ha4/GLos0yonQWYjIxutmIi5GQlagb+xBCzFC3YtnmX7i6sGJHwLzkEPTxP5hsSON2t8EBdaKJqKY/oTzviH8+88mcPNDjOi7PWnKck5qunsnZO1DDOzrkE5zo='
    }
    header = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded',
        'Host': '192.168.100.179',
        'Origin': 'http://192.168.100.179',
        'Refer': 'http://192.168.100.179/HTMSOFT/TasksQuery.aspx',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36'
    }
    

    data = parse.urlencode(data).encode()
    req = request.Request(url=url, data=data, headers=header)
    handler = request.HTTPCookieProcessor(cookie)
    opener = request.build_opener(handler)
    page = opener.open(req).read()

    soup = BeautifulSoup(page.decode(), 'html.parser')

    # print(page.decode())
    tag = soup.find("tr", {"class": "Grid_Item"})
    ret = []
    try:
        for td in tag.find_all("td"):
            ret.append(td.string)
    except:
        pass
    return tuple(ret)



# 测试SDM数据抓取方法是否可行
login("符杰", "123")
print(fetchInfo("6070003110"))
'''
q = Queue()
req_finnished = 0

login("符杰", "123")
def getInfoPutoToList(order):
    global req_finnished
    q.put(fetchInfo(order))
    req_finnished += 1
def fetchInfoMuliThreading(order_list):
    for order in order_list:
        _t = threading.Thread(target=getInfoPutoToList, args=(order,))
        _t.start()



# 从文件获取所有订单
login("符杰", "123")
xlsx_path = "./原3月底前计划要发的机器汇总.xlsx"
wb = openpyxl.load_workbook(xlsx_path)
output_wb = openpyxl.Workbook()
output_ws = wb.worksheets[0]
order_list = []
for ws in (wb["合同机（已出仓整机装配）"], wb["未出仓机器"]):
    for row in range(1, ws.max_row + 1):
        value = str(ws.cell(row, 4).value)
        if value is None or value == '':
            continue
        res = re.findall(r"\d+", value)
        for order in res:
            print(order)
            order_list.append(order)
print(len(order_list))
# 测试
# order_list = ["6070001637", "6070002508", "6070002508", "6070002508"]

# 写入
fetchInfoMuliThreading(order_list)
wb = openpyxl.Workbook()
ws = wb.worksheets[0]
while True:
    try:
        if req_finnished >= len(order_list):
            print("完成", req_finnished)
            break
        info = q.get(timeout=60)
        print(info)
        row = ws.max_row + 1
        for i in range(len(info) - 1):
            ws.cell(row, i + 1).value = info[i]
    except Exception as e:
        print("超时")
        break
wb.save("./SDM机器信息.xlsx")

'''